import openpyxl

class Cases:
    '''专门存储我们的测试数据'''
    def __init__(self):
        self.case_id = None
        self.title = None
        self.url = None
        self.data = None
        self.method = None
        self.expected = None
        self.actual = None
        self.result = None

class DoExcel:
     def __init__(self,file_name):
         #操作的文件
         self.file_name = file_name

         #实例化一个workbook对象
         #异常处理如何做
         try:
            self.workbook = openpyxl.load_workbook(filename=self.file_name)
         except Exception as e:
            print("错误：",e)

     def read(self,sheet_name):#读数据
         sheet = self.workbook[sheet_name]  #读取sheet
         cases=[]#存储数据
         max_row = sheet.max_row
         for i in range(2,max_row+1):#行的范围从第二行开始
             row_case = Cases()#每一行数据存在这个对象里面Cases() 对象/实例
             row_case.case_id = sheet.cell(row = i,column = 1).value # 存的是case_id
             row_case.title = sheet.cell(row = i,column = 2).value # 存的是title
             row_case.url = sheet.cell(row = i,column = 3).value # 存的是url
             row_case.data = sheet.cell(row = i,column = 4).value # 存的是data
             row_case.method = sheet.cell(row = i,column = 5).value # 存的是method
             row_case.expected = sheet.cell(row = i,column = 6).value # 存的是expected

             cases.append(row_case)

         return cases # for循环结束后返回cases列表

     def write_back(self,sheet_name,row,actual,result):#写回数据
        sheet=self.workbook[sheet_name]
        sheet.cell(row,7).value=actual#写入实际结果到第7列单元格里面去
        sheet.cell(row,8).value=result#写入执行结果到第8列单元格里面去
        self.workbook.save(self.file_name)#要记得保存，同时要记得Excel要关闭状态

if __name__=='__main__':
    from common import c02_contants
    from common.c03_request import Request
    #test_data=DoExcel('..//datas//testcase.xlsx').read('login')
    do_excel=DoExcel(c02_contants.case_file)
    test_data=do_excel.read('login')
    print(test_data)

    # #Test_data=DoExcel('..//datas//testcase.xlsx').write_back('login',2,'hao','PASS')

    request=Request() # 实例化对象
    for case in test_data:
        #参数的处理
        #print(type(case.data),case.data)
        resp= request.request1(case.method,case.url,case.data)
        # print(resp.json()) # 把Excel表中的data数据转成字典  ，
        # 在request模块里面在request1方法里面将字符串转化为字典，令每一个不同的case都直接已转换成字典
        if resp.text==case.expected:
            do_excel.write_back('login',case.case_id+1,resp.text,'PASS')
            print("第{0}条用例执行结果：PASS".format(case.case_id))
        else:
            do_excel.write_back('login',case.case_id+1,resp.text,'FAILED')
            print("第{0}条用例执行结果：Failed".format(case.case_id))


